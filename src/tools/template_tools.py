from __future__ import annotations

import re
import base64
from pathlib import Path
from typing import Dict, Any, List, Optional
from pydantic import BaseModel, Field
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from components.template_config.mapping_resolver import MappingResolver
from generators.template_generator import TemplateDocumentGenerator, mime_for
from utils.template_utils import (
    create_temp_template_from_base64,
    safe_render_template,
    format_mapping_for_display,
    validate_template_placeholders
)

_PLACEHOLDER_RE = re.compile(r"\{\{\s*([a-zA-Z_][a-zA-Z0-9_.]*)\s*\}\}", re.IGNORECASE)
_TABLE_PLACEHOLDER_RE = re.compile(r"^\s*\{\{\s*/([a-zA-Z_][a-zA-Z0-9_.]*)\s*\}\}\s*$")


class GenerateFromBase64TemplateReq(BaseModel):
    base64_template: str = Field(..., description="Base64-encoded .docx or .xlsx template")
    data: Dict[str, Any] = Field(..., description="Source JSON data")

    # Optional
    mapping: Optional[List[Dict[str, Any]]] = Field(default=None, description="Optional explicit mapping rules")
    output_filename: Optional[str] = Field(default=None, description="Custom output filename")
    output_format: Optional[str] = Field(default=None, description="Optional: docx, xlsx, pdf")
    auto_fallback: bool = Field(default=True, description="Fallback to direct key match if path fails")
    validate_template: bool = Field(default=True, description="Validate placeholders before rendering")
    table_mappings: Optional[List[Dict[str, Any]]] = Field(default=None, description="Table column mappings")
    table_placeholder: Optional[str] = Field(default=None, description="Table placeholder name")


class GenerateFromBase64TemplateOutput(BaseModel):
    type: str = "file_base64"
    filename: str
    mime_type: str
    base64: str
    size_bytes: int
    path: str
    meta: Dict[str, Any] = Field(default_factory=dict)
    missing_vars: List[str] = Field(default_factory=list)


def extract_template_structure(template_path: Path) -> tuple[List[str], Dict[str, List[Dict[str, str]]]]:
    suffix = template_path.suffix.lower()
    placeholders = set()
    table_mappings: Dict[str, List[Dict[str, str]]] = {}

    if suffix == ".docx":
        try:
            tpl = DocxTemplate(str(template_path))
            placeholders.update(tpl.get_undeclared_template_variables())
        except Exception:
            pass

    if suffix == ".xlsx":
        try:
            wb = load_workbook(str(template_path), data_only=False)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Detect tables
                table_mappings.update(_detect_excel_tables(ws))
                # Detect scalar placeholders
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            placeholders.update(_PLACEHOLDER_RE.findall(cell.value))

        except Exception:
            pass

    return sorted(list(placeholders)), table_mappings


def _detect_excel_tables(ws) -> Dict[str, List[Dict[str, str]]]:
    table_mappings: Dict[str, List[Dict[str, str]]] = {}

    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if not isinstance(cell.value, str):
                continue

            match = _TABLE_PLACEHOLDER_RE.match(cell.value)
            if not match:
                continue

            table_name = match.group(1)
            header_row_idx = row_idx - 1
            if header_row_idx < 1:
                continue

            # Extract headers
            headers: List[str] = []
            c = col_idx
            while c <= ws.max_column:
                h_cell = ws.cell(row=header_row_idx, column=c)
                h_val = str(h_cell.value).strip() if h_cell.value else ""
                if not h_val:
                    break
                headers.append(h_val)
                c += 1

            if headers:
                table_mappings[table_name] = [
                    {"column": header, "key": re.sub(r"[^a-zA-Z0-9_]", "_", header.lower())}
                    for header in headers
                ]

    return table_mappings


class TemplateTools:
    def __init__(self, template_generator: TemplateDocumentGenerator):
        self.tg = template_generator

    def generate_from_template(self, req: GenerateFromBase64TemplateReq) -> GenerateFromBase64TemplateOutput:
        """
        Generate Base64 Document from Base64 Template.
        """
        template_path = create_temp_template_from_base64(req.base64_template)

        try:
            raw_placeholders, detected_table_mappings = extract_template_structure(template_path)
            meta_placeholders = validate_template_placeholders(raw_placeholders)

            missing_vars: List[str] = []
            if req.validate_template and template_path.suffix.lower() == ".docx":
                def error_handler(ph: str, msg: str):
                    if not MappingResolver.is_table_placeholder(ph):  # skip tables
                        missing_vars.append(ph)

                safe_render_template(template_path, {}, error_handler)

            # Table logic
            table_placeholder = req.table_placeholder
            if table_placeholder is None and len(detected_table_mappings) == 1:
                table_placeholder = next(iter(detected_table_mappings.keys()))

            effective_table_mappings = req.table_mappings
            if effective_table_mappings is None and table_placeholder and table_placeholder in detected_table_mappings:
                effective_table_mappings = detected_table_mappings[table_placeholder]

            # Separate scalar vs table placeholders
            scalar_placeholders = [ph for ph in meta_placeholders if not MappingResolver.is_table_placeholder(ph)]
            auto_mappings = [{"placeholder": ph, "path": ph} for ph in scalar_placeholders]
            effective_mappings = req.mapping or auto_mappings

            context = MappingResolver.build_context_from_dict(
                data=req.data,
                mapping_dict=effective_mappings,
                table_mappings=effective_table_mappings,
                table_placeholder=table_placeholder,
                auto_fallback=req.auto_fallback
            )

            if table_placeholder and effective_table_mappings:
                incoming_rows = req.data.get(table_placeholder, [])
                if isinstance(incoming_rows, list):
                    normalized_rows = []
                    for row in incoming_rows:
                        row = row or {}
                        out_row = {}
                        for m in effective_table_mappings:
                            header = m.get("column")
                            key = m.get("key")
                            if not header:
                                continue
                            val = None
                            if key and isinstance(row, dict):
                                val = row.get(key)
                            if val is None and isinstance(row, dict):
                                val = row.get(header)
                                if val is None:
                                    val = row.get(header.lower())
                            out_row[header] = "" if val is None else val
                        normalized_rows.append(out_row)
                    context[table_placeholder] = normalized_rows

            filepath = self.tg.generate(
                template_path=template_path,
                context=context,
                output_filename=req.output_filename,
                output_format=req.output_format
            )

            out_path = Path(filepath)
            file_bytes = out_path.read_bytes()

            used_mapping = "Auto-generated from scalar placeholders"
            if req.mapping:
                used_mapping = format_mapping_for_display(req.mapping)
            if effective_table_mappings and table_placeholder:
                used_mapping += f" | table '{table_placeholder}' cols={len(effective_table_mappings)}"

            meta = {
                "used_mapping": used_mapping,
                "placeholders": meta_placeholders,
                "detected_tables": detected_table_mappings,
            }

            return GenerateFromBase64TemplateOutput(
                filename=out_path.name,
                path=str(out_path),
                mime_type=mime_for(out_path.suffix.lstrip(".").lower()),
                base64=base64.b64encode(file_bytes).decode("utf-8"),
                size_bytes=len(file_bytes),
                missing_vars=missing_vars,
                meta=meta
            )

        except Exception as e:
            raise ValueError(f"Generation failed: {e}")
        finally:
            if template_path.exists():
                template_path.unlink()