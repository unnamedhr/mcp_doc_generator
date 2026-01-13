from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from components.templates.excel import ExcelComponentsBase, ExcelPresets, ExcelStyles


OUTPUT_DIR = Path("./generated_documents")
OUTPUT_DIR.mkdir(exist_ok=True)


def _safe_json_loads(value: Any, default):
    if value is None:
        return default
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return default
        return json.loads(s)
    return default


def _hex_to_rgb(hex_color: Optional[str], default: str = "FFFFFF") -> str:
    if not hex_color:
        return default
    s = str(hex_color).strip()
    if s.startswith("#"):
        s = s[1:]
    s = s.upper()
    return s if re.fullmatch(r"[0-9A-F]{6}", s) else default


def _normalize_number_format(fmt: Optional[str]) -> Optional[str]:
    if not fmt:
        return None

    s = str(fmt).strip()
    # If it's already a classic Excel format string, keep it.
    if any(ch in s for ch in ["#", "0", "%", "€", "$"]) or s.lower() in ["general"]:
        # But avoid passing through "number:0" etc.
        if ":" not in s:
            return s

    if s.lower().startswith("number:"):
        decimals = s.split(":", 1)[1].strip()
        if decimals.isdigit():
            d = int(decimals)
            return "0" if d == 0 else "0." + ("0" * d)

    if s.lower().startswith("percent:"):
        decimals = s.split(":", 1)[1].strip()
        if decimals.isdigit():
            d = int(decimals)
            return "0%" if d == 0 else "0." + ("0" * d) + "%"

    if s.lower() in ["currency:euro", "currency:eur"]:
        return "€ #,##0.00"

    if s.lower() in ["currency:usd", "currency:dollar"]:
        return "$#,##0.00"

    return None


def _border_from(border_cfg: Dict[str, Any], style_override: Optional[str] = None) -> Border:
    style = style_override or border_cfg.get("style", "thin")
    color = _hex_to_rgb(border_cfg.get("color", "#000000"), default="000000")

    def side(enabled: bool) -> Side:
        return Side(style=style, color=color) if enabled else Side(style=None)

    return Border(
        left=side(bool(border_cfg.get("left", True))),
        right=side(bool(border_cfg.get("right", True))),
        top=side(bool(border_cfg.get("top", True))),
        bottom=side(bool(border_cfg.get("bottom", True))),
    )


def _apply_cell_format(cell, cfg: Dict[str, Any], base_border_cfg: Dict[str, Any]):
    font_color = cfg.get("font_color", cfg.get("text_color", cfg.get("color")))
    bg_color = cfg.get("bg_color", cfg.get("cell_color"))
    border_style = cfg.get("border_style")

    # Font
    cell.font = Font(
        name=cfg.get("font_name", "Calibri"),
        size=cfg.get("font_size", 10),
        bold=bool(cfg.get("bold", False)),
        italic=bool(cfg.get("italic", False)),
        underline=("single" if cfg.get("underline") is True else (cfg.get("underline") if cfg.get("underline") else None)),
        color=_hex_to_rgb(font_color, default="000000") if font_color else None,
    )

    # Fill
    if bg_color:
        cell.fill = PatternFill(
            start_color=_hex_to_rgb(bg_color, default="FFFFFF"),
            end_color=_hex_to_rgb(bg_color, default="FFFFFF"),
            fill_type="solid",
        )

    # Alignment
    cell.alignment = Alignment(
        horizontal=cfg.get("alignment", "center"),
        vertical=cfg.get("vertical_alignment", "center"),
        wrap_text=bool(cfg.get("wrap_text", False)),
    )

    # Border
    cell.border = _border_from(base_border_cfg, style_override=border_style)

    # Number format
    nf = cfg.get("number_format")
    nf = _normalize_number_format(nf) or nf
    if nf:
        cell.number_format = nf


def _set_row_height(ws, row_idx: int, height: Optional[float]):
    if height is None:
        return
    try:
        ws.row_dimensions[row_idx].height = float(height)
    except (TypeError, ValueError):
        return


def _cell_range_for_table(num_cols: int, num_rows: int) -> str:
    # includes the header row
    last_col = get_column_letter(num_cols)
    last_row = num_rows
    return f"A1:{last_col}{last_row}"


def generate_excel_report(
    title: str,
    sheet_name: str,
    headers: str,
    data_rows: str,
    styling_config: str = "{}",
    sheet_config: str = "{}",
    include_totals: bool = False,
    totals_config: str = "{}",
    include_data_validation: bool = False,
    validation_config: str = "{}",
    include_freeze_panes: bool = True,
    freeze_config: str = "{}",
    include_autofilter: bool = False,
    filter_config: str = "{}",
    include_conditional_formatting: bool = False,
    conditional_config: str = "{}",
    columns_config: str = "{}",
    preset_theme: str = "",
) -> str:
    try:
        headers_list: List[str] = _safe_json_loads(headers, default=[])
        data_list: List[List[Any]] = _safe_json_loads(data_rows, default=[])
        columns_cfg: Dict[str, Any] = _safe_json_loads(columns_config, default={})

        raw_sheet_cfg: Dict[str, Any] = _safe_json_loads(sheet_config, default={})
        raw_totals_cfg: Dict[str, Any] = _safe_json_loads(totals_config, default={})
        raw_validation_cfg: Dict[str, Any] = _safe_json_loads(validation_config, default={})
        raw_freeze_cfg: Dict[str, Any] = _safe_json_loads(freeze_config, default={})
        raw_filter_cfg: Dict[str, Any] = _safe_json_loads(filter_config, default={})
        raw_conditional_cfg: Dict[str, Any] = _safe_json_loads(conditional_config, default={})

        # Styles
        if preset_theme:
            preset_method = getattr(ExcelPresets, preset_theme, None)
            style_config = preset_method() if callable(preset_method) else _safe_json_loads(styling_config, default={})
        else:
            style_config = _safe_json_loads(styling_config, default={})

        header_cfg = ExcelComponentsBase.render_header(style_config.get("header", {}))
        data_cfg = ExcelComponentsBase.render_data_rows(style_config.get("data", {}))

        base_border_cfg = ExcelStyles.get_border_config(style_config.get("borders", {}))

        if include_totals:
            raw_totals_cfg.setdefault("enabled", True)
        totals_cfg_final = ExcelComponentsBase.render_totals_row(raw_totals_cfg)

        if include_freeze_panes:
            raw_freeze_cfg.setdefault("enabled", True)
        freeze_final = ExcelComponentsBase.get_freeze_panes(raw_freeze_cfg)

        if include_autofilter:
            raw_filter_cfg.setdefault("enabled", True)
        filter_final = ExcelComponentsBase.get_table_autofilter(raw_filter_cfg)

        if include_data_validation:
            raw_validation_cfg.setdefault("enabled", True)
        validation_final = ExcelComponentsBase.get_data_validation(raw_validation_cfg)

        if include_conditional_formatting:
            raw_conditional_cfg.setdefault("enabled", True)
        conditional_final = ExcelComponentsBase.get_conditional_formatting(raw_conditional_cfg)

        # Sheet config via template
        sheet_final = ExcelComponentsBase.render_sheet({
            **raw_sheet_cfg,
            "name": raw_sheet_cfg.get("name", sheet_name),
        })

        # Create workbook/sheet
        wb = Workbook()
        ws = wb.active
        ws.title = (sheet_final.get("name") or sheet_name)[:31]

        # Apply sheet state and tab colour
        state = sheet_final.get("state")
        if state in ("visible", "hidden", "veryHidden"):
            ws.sheet_state = state
        tab_color = sheet_final.get("tab_color")
        if tab_color:
            ws.sheet_properties.tabColor = _hex_to_rgb(tab_color, default="0070C0")

        # -------------------
        # Header row
        # -------------------
        if header_cfg.get("enabled", True):
            _set_row_height(ws, 1, header_cfg.get("height"))
            for col_num, header_text in enumerate(headers_list, 1):
                cell = ws.cell(row=1, column=col_num, value=header_text)
                _apply_cell_format(cell, header_cfg, base_border_cfg)
                col_letter = get_column_letter(col_num)
                col_override = columns_cfg.get(col_letter, {}) or {}
                if col_override:
                    merged = {**header_cfg, **col_override}
                    _apply_cell_format(cell, merged, base_border_cfg)

        # -------------------
        # Data rows
        # -------------------
        start_data_row = 2
        if data_cfg.get("enabled", True):
            for i, row_data in enumerate(data_list):
                row_num = start_data_row + i
                _set_row_height(ws, row_num, data_cfg.get("height"))

                is_alternating = (i % 2 == 1)
                for col_num, value in enumerate(row_data, 1):
                    # Ensure formulas are stored as formulas (must start with "=")
                    if isinstance(value, str):
                        v = value.strip()
                        if v.startswith("="):
                            value = v

                    cell = ws.cell(row=row_num, column=col_num, value=value)

                    fmt = dict(data_cfg)
                    if fmt.get("alternating_colors", True) and is_alternating:
                        fmt["bg_color"] = fmt.get("alternate_color", "#F7FAFC")

                    _apply_cell_format(cell, fmt, base_border_cfg)

                    col_letter = get_column_letter(col_num)
                    col_override = columns_cfg.get(col_letter, {}) or {}
                    if col_override:
                        merged = {**fmt, **col_override}
                        _apply_cell_format(cell, merged, base_border_cfg)
                        if "format" in col_override and not merged.get("number_format"):
                            nf = _normalize_number_format(col_override.get("format"))
                            if nf:
                                cell.number_format = nf

        # -------------------
        # Totals row
        # -------------------
        totals_row_idx = None
        if include_totals and totals_cfg_final.get("enabled", False) and data_list:
            totals_row_idx = start_data_row + len(data_list)
            _set_row_height(ws, totals_row_idx, totals_cfg_final.get("height"))

            label = totals_cfg_final.get("label", "TOTAL")
            label_cell = ws.cell(row=totals_row_idx, column=1, value=label)
            _apply_cell_format(label_cell, totals_cfg_final, base_border_cfg)

            formula_type = (totals_cfg_final.get("formula_type") or "SUM").upper()
            data_end_row = totals_row_idx - 1

            for col_num in range(2, len(headers_list) + 1):
                col_letter = get_column_letter(col_num)
                if formula_type == "AVERAGE":
                    formula = f"=AVERAGE({col_letter}{start_data_row}:{col_letter}{data_end_row})"
                elif formula_type == "COUNT":
                    formula = f"=COUNT({col_letter}{start_data_row}:{col_letter}{data_end_row})"
                elif formula_type == "MIN":
                    formula = f"=MIN({col_letter}{start_data_row}:{col_letter}{data_end_row})"
                elif formula_type == "MAX":
                    formula = f"=MAX({col_letter}{start_data_row}:{col_letter}{data_end_row})"
                else:
                    formula = f"=SUM({col_letter}{start_data_row}:{col_letter}{data_end_row})"

                c = ws.cell(row=totals_row_idx, column=col_num, value=formula)
                _apply_cell_format(c, totals_cfg_final, base_border_cfg)

        # -------------------
        # Column widths / formats
        # -------------------
        for col_num in range(1, len(headers_list) + 1):
            col_letter = get_column_letter(col_num)
            col_cfg = columns_cfg.get(col_letter, {}) or {}

            if "width" in col_cfg and col_cfg["width"]:
                ws.column_dimensions[col_letter].width = float(col_cfg["width"])
            else:
                # Default expected width
                ws.column_dimensions[col_letter].width = 15

        # -------------------
        # Freeze panes
        # -------------------
        if include_freeze_panes and freeze_final.get("enabled", False):
            freeze_cell = freeze_final.get("freeze_cell") or "A2"
            ws.freeze_panes = freeze_cell

        # -------------------
        # AutoFilter
        # -------------------
        if include_autofilter and filter_final.get("enabled", False):
            total_rows = 1 + len(data_list)
            if totals_row_idx:
                total_rows += 1
            ref = filter_final.get("range") or _cell_range_for_table(len(headers_list), total_rows)
            ws.auto_filter.ref = ref

        # -------------------
        # Data validation
        # -------------------
        if include_data_validation and validation_final.get("enabled", False):
            dv_type = validation_final.get("type", "list")
            dv_formula = validation_final.get("formula", '"Option1,Option2,Option3"')
            dv_operator = validation_final.get("operator", None)
            dv_allow_blank = bool(validation_final.get("allow_blank", True))

            dv = DataValidation(
                type=dv_type,
                formula1=dv_formula,
                operator=dv_operator,
                allow_blank=dv_allow_blank,
                showErrorMessage=True,
                showInputMessage=bool(validation_final.get("show_input", False)),
            )
            dv.errorTitle = validation_final.get("error_title", "Invalid Entry")
            dv.error = validation_final.get("error_message", "Please enter a valid value")

            if validation_final.get("input_title"):
                dv.promptTitle = validation_final["input_title"]
            if validation_final.get("input_message"):
                dv.prompt = validation_final["input_message"]

            ws.add_data_validation(dv)
            dv_range = validation_final.get("range") or f"B2:B{len(data_list) + 1}"
            dv.add(dv_range)

        # -------------------
        # Conditional formatting (color scale)
        # -------------------
        if include_conditional_formatting and conditional_final.get("enabled", False):
            cond_type = conditional_final.get("type", "colorScale")
            cond_range = conditional_final.get("range") or f"B2:{get_column_letter(len(headers_list))}{len(data_list) + 1}"

            if cond_type == "colorScale":
                color_scale = ColorScaleRule(
                    start_type="min",
                    start_color=_hex_to_rgb(conditional_final.get("color_min", "#FFEB9C"), default="FFEB9C"),
                    mid_type="percentile",
                    mid_value=50,
                    mid_color=_hex_to_rgb(conditional_final.get("color_mid", "#FFFFFF"), default="FFFFFF"),
                    end_type="max",
                    end_color=_hex_to_rgb(conditional_final.get("color_max", "#92D050"), default="92D050"),
                )
                ws.conditional_formatting.add(cond_range, color_scale)

        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{title.replace(' ', '_')}_{timestamp}.xlsx"
        filepath = OUTPUT_DIR / filename
        wb.save(str(filepath))

        return f"✓ Excel generated: {filepath} | Rows: {len(data_list)}, Cols: {len(headers_list)}"

    except json.JSONDecodeError as e:
        return f"✗ JSON Error: {str(e)}"
    except Exception as e:
        return f"✗ Error generating Excel: {str(e)}"