from __future__ import annotations

import base64
import json
import re
from pathlib import Path
from typing import Dict, Any, Optional, List, TypedDict, Union
from datetime import datetime

from docxtpl import DocxTemplate
from openpyxl import load_workbook

_PLACEHOLDER_RE = re.compile(r"\{\{\s*([a-zA-Z_][a-zA-Z0-9_.]*)\s*}}")


class TemplateResult(TypedDict):
    type: str
    filename: str
    mime_type: str
    base64: str
    size_bytes: int
    path: str
    meta: Dict[str, Any]


def _file_to_base64(path: Path) -> str:
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def _mime_for(fmt: str) -> str:
    fmt = fmt.lower()
    if fmt == "docx":
        return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    if fmt == "xlsx":
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if fmt == "pdf":
        return "application/pdf"
    return "application/octet-stream"

class TemplateDocumentGenerator:
    """Generate documents from .docx and .xlsx templates."""

    def __init__(self, output_dir: Union[str, Path]):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)


    def generate(
        self,
        template_path: Path,
        context: Dict[str, Any],
        output_filename: Optional[str] = None,
        output_format: Optional[str] = None,
    ) -> str:
        """
        Generate a document from a template.
        """
        template_path = Path(template_path)

        if isinstance(context, str):
            context_dict = json.loads(context) if context.strip() else {}
        else:
            context_dict = context or {}

        if not template_path.exists():
            raise ValueError(f"Template not found: {template_path}")

        template_suffix = template_path.suffix.lower()
        fmt = (output_format or template_suffix.lstrip(".")).lower()

        if fmt not in ("docx", "xlsx", "pdf"):
            raise ValueError("output_format must be one of: docx, xlsx, pdf")

        # Validate template-vs-format compatibility
        if fmt in ("docx", "pdf") and template_suffix != ".docx":
            raise ValueError(f"output_format='{fmt}' requires a .docx template, got {template_suffix}")
        if fmt == "xlsx" and template_suffix != ".xlsx":
            raise ValueError(f"output_format='xlsx' requires a .xlsx template, got {template_suffix}")

        # Filename
        if not output_filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ext = "pdf" if fmt == "pdf" else fmt
            output_filename = f"generated_{timestamp}.{ext}"

        output_path = self.output_dir / output_filename

        # Filename
        if not output_filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ext = "pdf" if fmt == "pdf" else fmt
            output_filename = f"generated_{timestamp}.{ext}"

        output_path = self.output_dir / output_filename

        # Generate
        if fmt == "docx":
            self._generate_docx(template_path, context, output_path)
            return str(output_path)

        if fmt == "xlsx":
            self._generate_xlsx(template_path, context, output_path)
            return str(output_path)

        # Generate a temporary docx next to the output PDF, then convert.
        tmp_docx_path = output_path.with_suffix(".docx")
        self._generate_docx(template_path, context, tmp_docx_path)

        pdf_path = self.convert_to_pdf(tmp_docx_path)
        if not pdf_path:
            raise ValueError("DOCX to PDF conversion failed")
        return pdf_path

    def generate_base64(
            self,
            template_path: Union[str, Path],
            context: Union[Dict[str, Any], str],
            output_filename: Optional[str] = None,
            output_format: Optional[str] = None,
            max_bytes: int = 8 * 1024 * 1024,
    ) -> TemplateResult:
        """
        MCP-friendly: generate and return base64 and metadata.
        """
        out_path_str = self.generate(
            template_path=template_path,
            context=context,
            output_filename=output_filename,
            output_format=output_format,
        )

        out_path = Path(out_path_str)
        if not out_path.exists():
            raise RuntimeError(f"Template generation succeeded but output not found: {out_path}")

        file_bytes = out_path.read_bytes()

        if len(file_bytes) > max_bytes:
            raise RuntimeError(
                f"Generated file is too large ({len(file_bytes)} bytes). Reduce template expansion or data."
            )

        fmt = out_path.suffix.lstrip(".").lower()

        return {
            "type": "file_base64",
            "filename": out_path.name,
            "path": str(out_path),
            "mime_type": _mime_for(fmt),
            "base64": base64.b64encode(file_bytes).decode("utf-8"),
            "size_bytes": len(file_bytes),
            "meta": {
                "format": fmt,
                "template": str(Path(template_path)),
                "timestamp": datetime.now().strftime("%Y%m%d_%H%M%S"),
            },
        }


    @staticmethod
    def _generate_docx(template_path: Path, context: Dict[str, Any], output_path: Path) -> None:
        try:
            tpl = DocxTemplate(str(template_path))
        except Exception as e:
            raise ValueError(f"Failed to load DOCX template: {e}")

        try:
            tpl.render(context)
        except Exception as e:
            raise ValueError(f"Jinja2 render error: {e}")

        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            tpl.save(str(output_path))
        except Exception as e:
            raise ValueError(f"Failed to save DOCX document: {e}")


    def _generate_xlsx(self, template_path: Path, context: Dict[str, Any], output_path: Path) -> None:
        try:
            wb = load_workbook(str(template_path))
        except Exception as e:
            raise ValueError(f"Failed to load XLSX template: {e}")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self._replace_scalars(ws, context)
            self._expand_table_placeholder(ws, context)

        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(output_path))
        except Exception as e:
            raise ValueError(f"Failed to save XLSX document: {e}")


    @staticmethod
    def _replace_scalars(ws, context: Dict[str, Any]) -> None:
        table_placeholder_re = re.compile(r"\{\{\s*/[a-zA-Z_][a-zA-Z0-9_.]*\s*\}\}")

        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.strip():
                    continue

                # Skip cells containing table placeholders
                if table_placeholder_re.search(cell.value):
                    continue

                def replacer(match):
                    key = match.group(1)
                    val = context.get(key)
                    return "" if val is None else str(val)

                cell.value = _PLACEHOLDER_RE.sub(replacer, cell.value)


    @staticmethod
    def _expand_table_placeholder(ws, context: Dict[str, Any]) -> None:
        table_placeholder_re = re.compile(r"\{\{\s*/([a-zA-Z_][a-zA-Z0-9_.]*)\s*\}\}")

        markers = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    match = table_placeholder_re.search(cell.value)
                    if match:
                        table_key = match.group(1)
                        markers.append((cell.row, cell.column, table_key))

        for start_row, start_col, table_key in markers:
            table_rows: List[Dict[str, Any]] = context.get(table_key) or []
            if not isinstance(table_rows, list) or not table_rows:
                continue

            header_row = start_row - 1
            if header_row < 1:
                continue

            headers = []
            col = start_col
            while col <= ws.max_column:
                header_val = ws.cell(row=header_row, column=col).value
                if header_val is None or str(header_val).strip() == "":
                    break
                headers.append(str(header_val).strip())
                col += 1

            # Clear the table placeholder cell
            ws.cell(row=start_row, column=start_col).value = None

            # Fill table rows
            for i, row_data in enumerate(table_rows):
                r = start_row + i
                for j, header in enumerate(headers):
                    c = start_col + j
                    ws.cell(row=r, column=c).value = row_data.get(header, "")


    def convert_to_pdf(self, docx_path: Path) -> Optional[str]:
        """Convert DOCX to PDF using LibreOffice."""
        import subprocess

        docx_path = Path(docx_path)
        if not docx_path.exists():
            raise ValueError(f"File not found: {docx_path}")

        try:
            result = subprocess.run(
                [
                    "libreoffice",
                    "--headless",
                    "--convert-to", "pdf",
                    "--outdir", str(self.output_dir),
                    str(docx_path)
                ],
                capture_output=True,
                timeout=30
            )

            if result.returncode != 0:
                return None

            pdf_path = self.output_dir / f"{docx_path.stem}.pdf"
            return str(pdf_path) if pdf_path.exists() else None

        except Exception:
            return None